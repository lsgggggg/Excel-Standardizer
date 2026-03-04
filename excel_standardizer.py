#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel Standardizer v3.0 — 全面重构版
=====================================
核心原则:
  1. 宁可漏改，不可误改 — 默认只启用不可能出错的规则
  2. 公式绝不触碰 — 检测到公式直接跳过
  3. 格式绝不覆盖 — 绝不修改原始样式/字体/颜色
  4. 先预览后执行 — 两阶段处理: 扫描→审核→应用
  5. 列感知处理 — 公司名规则只对公司名列, 邮箱规则只对邮箱列
  6. 精度不可损失 — 数字类型保持原样, 不做string round-trip
"""

import os
import sys
import re
import json
import copy
import unicodedata
import html as html_module
from datetime import datetime, date, time as dt_time
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("错误: 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

try:
    from opencc import OpenCC
    HAS_OPENCC = True
except ImportError:
    HAS_OPENCC = False

try:
    from unidecode import unidecode
    HAS_UNIDECODE = True
except ImportError:
    HAS_UNIDECODE = False


# ============================================================
#  安全等级定义 (SAFETY LEVELS)
# ============================================================

SAFETY_LEVELS = {
    'SAFE': [
        "1.03_nbsp_to_space", "1.06_collapse_spaces", "1.07_strip_whitespace",
        "1.08_zero_width_chars", "1.09_control_chars", "1.11_soft_hyphen",
        "1.12_bidi_marks", "8.01_strip_cell", "8.02_collapse_cell_spaces",
        "8.07_html_entities",
    ],
    'MODERATE': [
        "1.01_fullwidth_to_halfwidth", "1.02_fullwidth_space", "1.04_tab_to_space",
        "1.05_newline_to_space", "1.10_unicode_normalize",
        "4.02_tilde_normalize", "4.05_ellipsis_normalize",
        "5.06_negative_accounting", "8.09_email_normalize",
        "11.01_header_clean",
    ],
    'DANGEROUS': [
        "2.01_case_transform", "3.01_traditional_to_simplified",
        "3.02_chinese_punctuation", "3.03_chinese_brackets",
        "3.04_chinese_comma_dot", "3.05_simplified_to_traditional",
        "3.06_chinese_number_to_arabic", "3.07_chinese_financial_number",
        "3.08_chinese_region_normalize",
        "4.01_brackets_normalize", "4.03_quotes_normalize",
        "4.04_dash_normalize", "4.06_consecutive_punctuation",
        "4.07_special_symbols_clean",
        "5.02_thousand_separator", "5.03_currency_symbols",
        "5.04_percent_handling", "5.05_scientific_notation",
        "5.07_number_with_unit",
        "6.01_date_normalize", "6.02_time_normalize",
        "7.01_company_suffix_normalize", "7.02_company_intl_suffix",
        "7.03_chinese_company_type", "7.04_company_stop_words",
        "7.05_chinese_city_brackets",
        "8.03_spelling_variant", "8.04_abbreviation_mapping",
        "8.05_diacritics_remove", "8.06_emoji_remove",
        "8.08_url_normalize", "8.10_boolean_normalize",
        "8.11_null_normalize", "8.12_gender_normalize",
        "9.01_phone_normalize", "9.02_id_card_normalize",
        "9.03_credit_code_normalize", "9.04_bank_card_normalize",
        "12.01_address_abbreviation",
    ],
}

# 规则适用的列类型映射
RULE_COLUMN_APPLICABILITY = {
    "2.01_case_transform": {'email'},
    "7.01_company_suffix_normalize": {'company_name'},
    "7.02_company_intl_suffix": {'company_name'},
    "7.03_chinese_company_type": {'company_name'},
    "7.04_company_stop_words": {'company_name'},
    "7.05_chinese_city_brackets": {'company_name', 'address'},
    "8.09_email_normalize": {'email'},
    "8.10_boolean_normalize": {'boolean'},
    "8.12_gender_normalize": {'gender'},
    "9.01_phone_normalize": {'phone'},
    "9.02_id_card_normalize": {'id_number'},
    "9.03_credit_code_normalize": {'id_number'},
    "9.04_bank_card_normalize": {'id_number', 'bank_card'},
    "12.01_address_abbreviation": {'address'},
    "3.08_chinese_region_normalize": {'address'},
}


# ============================================================
#  默认设置 (SAFE DEFAULTS)
# ============================================================

DEFAULT_SETTINGS = {
    # ── 一、字符编码与不可见字符清理 ──
    "1.01_fullwidth_to_halfwidth": {"enabled": True, "desc": "全角字符→半角字符(排除日文片假名)"},
    "1.02_fullwidth_space": {"enabled": True, "desc": "全角空格(U+3000)→半角空格"},
    "1.03_nbsp_to_space": {"enabled": True, "desc": "不间断空格(\\xa0)→普通空格"},
    "1.04_tab_to_space": {"enabled": True, "desc": "Tab制表符→空格"},
    "1.05_newline_to_space": {"enabled": False, "desc": "换行符(\\r\\n)→空格(默认关闭,换行可能有意义)"},
    "1.06_collapse_spaces": {"enabled": True, "desc": "连续多个空格合并为1个"},
    "1.07_strip_whitespace": {"enabled": True, "desc": "首尾空白去除"},
    "1.08_zero_width_chars": {"enabled": True, "desc": "零宽字符清除"},
    "1.09_control_chars": {"enabled": True, "desc": "控制字符清除(U+0000-001F, U+007F)"},
    "1.10_unicode_normalize": {"enabled": True, "option": "NFC", "options": ["NFC", "NFKC", "NFD", "NFKD"],
                                "desc": "Unicode归一化(默认NFC,安全;NFKC会改变²→2等)"},
    "1.11_soft_hyphen": {"enabled": True, "desc": "软连字符(U+00AD)清除"},
    "1.12_bidi_marks": {"enabled": True, "desc": "双向文本标记清除"},

    # ── 二、大小写标准化 ──
    "2.01_case_transform": {"enabled": False, "option": "none",
                             "options": ["lower", "upper", "title", "none"],
                             "desc": "大小写转换(默认关闭!会破坏人名/缩写)"},

    # ── 三、中文相关标准化 ──
    "3.01_traditional_to_simplified": {"enabled": False, "desc": "繁体→简体(默认关闭,对港澳台数据有破坏性)"},
    "3.02_chinese_punctuation": {"enabled": False, "desc": "中文标点→英文标点(默认关闭,中文标点本身是正确的)"},
    "3.03_chinese_brackets": {"enabled": False, "desc": "中文全角括号→英文半角(仅圆括号,默认关闭)"},
    "3.04_chinese_comma_dot": {"enabled": False, "desc": "中文顿号→逗号(默认关闭,顿号有特定语义)"},
    "3.05_simplified_to_traditional": {"enabled": False, "desc": "简体→繁体(默认关闭)"},
    "3.06_chinese_number_to_arabic": {"enabled": False, "desc": "中文数字→阿拉伯数字(默认关闭,可能破坏地名)"},
    "3.07_chinese_financial_number": {"enabled": False, "desc": "大写金额→阿拉伯数字(默认关闭)"},
    "3.08_chinese_region_normalize": {"enabled": False, "desc": "省市区去后缀(默认关闭,可能造成歧义)"},

    # ── 四、标点符号与特殊字符 ──
    "4.01_brackets_normalize": {"enabled": False,
                                 "desc": "括号统一(默认关闭!仅转换中文圆括号,不动[]{}<>《》)"},
    "4.02_tilde_normalize": {"enabled": True, "desc": "波浪号统一 ～→~"},
    "4.03_quotes_normalize": {"enabled": False, "option": "double",
                               "options": ["double", "single"],
                               "desc": "引号统一(默认关闭,可能影响代码/CSV)"},
    "4.04_dash_normalize": {"enabled": False, "desc": "破折号统一为-(默认关闭,em-dash和hyphen有不同语义)"},
    "4.05_ellipsis_normalize": {"enabled": True, "option": "three_dots",
                                 "options": ["three_dots", "unicode_ellipsis"], "desc": "省略号统一"},
    "4.06_consecutive_punctuation": {"enabled": False, "desc": "连续标点去重(默认关闭)"},
    "4.07_special_symbols_clean": {"enabled": False, "desc": "清理™®©(默认关闭)"},

    # ── 五、数字与数值 ──
    "5.02_thousand_separator": {"enabled": False, "desc": "千分位清除(默认关闭,需确认数据为纯数字)"},
    "5.03_currency_symbols": {"enabled": False, "option": "remove",
                               "options": ["remove", "keep"],
                               "desc": "货币符号处理(默认关闭,会丢失币种信息)"},
    "5.04_percent_handling": {"enabled": False, "option": "keep_string",
                               "options": ["keep_string", "to_decimal"], "desc": "百分号处理(默认关闭)"},
    "5.05_scientific_notation": {"enabled": False, "desc": "科学计数法还原(默认关闭,大数有精度风险)"},
    "5.06_negative_accounting": {"enabled": True, "desc": "会计负数(100)→-100"},
    "5.07_number_with_unit": {"enabled": False, "desc": "带单位数字拆分(默认关闭)"},

    # ── 六、日期与时间 ──
    "6.01_date_normalize": {"enabled": False, "option": "YYYY-MM-DD",
                             "options": ["YYYY-MM-DD", "DD/MM/YYYY", "MM/DD/YYYY", "YYYY年MM月DD日"],
                             "desc": "日期格式统一(默认关闭)"},
    "6.02_time_normalize": {"enabled": False, "option": "HH:MM:SS",
                             "options": ["HH:MM:SS", "HH:MM", "12hour"], "desc": "时间格式统一(默认关闭)"},

    # ── 七、公司/机构名称(需列类型感知) ──
    "7.01_company_suffix_normalize": {"enabled": False, "option": "abbreviation",
                                       "options": ["abbreviation", "full_form"],
                                       "desc": "公司后缀统一(默认关闭,需指定公司名列)"},
    "7.02_company_intl_suffix": {"enabled": False, "desc": "国际公司后缀统一(默认关闭)"},
    "7.03_chinese_company_type": {"enabled": False, "desc": "中文公司类型统一(默认关闭)"},
    "7.04_company_stop_words": {"enabled": False, "desc": "公司停用词去除(默认关闭,极度危险)"},
    "7.05_chinese_city_brackets": {"enabled": False, "desc": "中文括号地名统一(默认关闭)"},

    # ── 八、通用文本 ──
    "8.01_strip_cell": {"enabled": True, "desc": "单元格前后空白去除"},
    "8.02_collapse_cell_spaces": {"enabled": True, "desc": "单元格内连续空格合并"},
    "8.03_spelling_variant": {"enabled": False, "option": "american",
                               "options": ["american", "british"], "desc": "拼写变体统一(默认关闭)"},
    "8.04_abbreviation_mapping": {"enabled": False, "desc": "缩写映射(默认关闭)"},
    "8.05_diacritics_remove": {"enabled": False, "desc": "变音符号清除(安全:自动跳过中日韩字符,默认关闭)"},
    "8.06_emoji_remove": {"enabled": False, "desc": "Emoji清除(默认关闭,已修正正则范围)"},
    "8.07_html_entities": {"enabled": True, "desc": "HTML实体转换(&amp;→&)"},
    "8.08_url_normalize": {"enabled": False, "option": "remove_protocol",
                            "options": ["remove_protocol", "extract_domain", "keep"],
                            "desc": "URL清理(默认关闭)"},
    "8.09_email_normalize": {"enabled": True, "desc": "邮箱格式标准化(转小写去空格)"},
    "8.10_boolean_normalize": {"enabled": False, "option": "yes_no",
                                "options": ["yes_no", "true_false", "1_0", "是_否"],
                                "desc": "布尔值统一(默认关闭)"},
    "8.11_null_normalize": {"enabled": False, "option": "empty",
                             "options": ["empty", "N/A", "NULL", "None", "-"],
                             "desc": "空值统一(默认关闭,'-'等可能有意义)"},
    "8.12_gender_normalize": {"enabled": False, "option": "M_F",
                               "options": ["M_F", "Male_Female", "男_女"],
                               "desc": "性别统一(默认关闭)"},

    # ── 九、电话号码与证件号 ──
    "9.01_phone_normalize": {"enabled": False, "desc": "电话号码标准化(默认关闭)"},
    "9.02_id_card_normalize": {"enabled": False, "desc": "身份证号X大写(默认关闭)"},
    "9.03_credit_code_normalize": {"enabled": False, "desc": "社会信用代码大写(默认关闭)"},
    "9.04_bank_card_normalize": {"enabled": False, "desc": "银行卡号清理(默认关闭)"},

    # ── 十、Excel格式(全部默认关闭,不动原始格式!) ──
    "10.01_freeze_header": {"enabled": False, "desc": "冻结首行(默认关闭,不改原始格式)"},
    "10.02_header_style": {"enabled": False, "desc": "表头样式(默认关闭,不改原始格式)"},
    "10.03_auto_column_width": {"enabled": False, "desc": "自动列宽(默认关闭,不改原始格式)"},
    "10.04_auto_filter": {"enabled": False, "desc": "自动筛选(默认关闭,不改原始格式)"},
    "10.05_number_format_unify": {"enabled": False, "desc": "数字格式统一(默认关闭)"},
    "10.06_font_unify": {"enabled": False, "option": "Arial 11",
                          "options": ["Arial 11", "微软雅黑 11", "Calibri 11"],
                          "desc": "统一字体(默认关闭,不改原始格式)"},

    # ── 十一、数据结构 ──
    "11.01_header_clean": {"enabled": True, "desc": "表头去空格换行"},
    "11.02_duplicate_rows": {"enabled": False, "option": "mark",
                              "options": ["mark", "remove"], "desc": "重复行检测(默认关闭)"},
    "11.03_empty_rows_clean": {"enabled": False, "desc": "空行清除(默认关闭)"},
    "11.04_empty_cols_clean": {"enabled": False, "desc": "空列清除(默认关闭)"},

    # ── 十二、地址 ──
    "12.01_address_abbreviation": {"enabled": False, "option": "expand",
                                    "options": ["expand", "abbreviate"],
                                    "desc": "地址缩写统一(默认关闭)"},
}

SETTINGS_FILE = "excel_standardizer_settings.json"


def load_settings():
    settings = copy.deepcopy(DEFAULT_SETTINGS)
    if os.path.exists(SETTINGS_FILE):
        try:
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                user_settings = json.load(f)
            for k, v in user_settings.items():
                if k in settings:
                    settings[k].update(v)
        except Exception:
            pass
    return settings


def save_settings(settings):
    with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)


# ============================================================
#  ChangeProposal — 变更提案
# ============================================================

class ChangeProposal:
    """一条拟变更提案，支持人工审核"""

    def __init__(self, sheet, cell_ref, row, col, original, proposed, rule_id, rule_desc):
        self.sheet = sheet
        self.cell_ref = cell_ref
        self.row = row
        self.col = col
        self.original = original
        self.proposed = proposed
        self.rule_id = rule_id
        self.rule_desc = rule_desc
        self.accepted = True
        self.user_override = None

    @property
    def final_value(self):
        if self.user_override is not None:
            return self.user_override
        return self.proposed if self.accepted else self.original


# ============================================================
#  ColumnTypeDetector — 列类型推断
# ============================================================

class ColumnTypeDetector:
    """根据表头关键词 + 数据采样推断列类型"""

    HEADER_KEYWORDS = {
        'person_name': ['姓名', '名字', '联系人', 'name', 'contact', 'person',
                        '收件人', '发件人', '负责人', '申请人', '经办人'],
        'company_name': ['公司', '企业', '机构', '单位', 'company', 'corp', 'firm',
                         'organization', 'org', '甲方', '乙方', '供应商', '客户名称'],
        'address': ['地址', '住址', 'address', 'addr', '所在地', '省', '市', '区',
                    '城市', 'city', 'state', 'province', 'region'],
        'phone': ['电话', '手机', '座机', 'phone', 'tel', 'mobile', 'fax', '联系电话'],
        'email': ['邮箱', '邮件', 'email', 'e-mail', 'mail'],
        'id_number': ['身份证', '证件号', '护照', 'id', 'passport', '社会信用代码',
                      '统一代码', '证照号', '工号', '学号'],
        'bank_card': ['银行卡', '卡号', 'card', 'account', '账号'],
        'currency_amount': ['金额', '价格', '费用', '成本', '收入', '支出', 'amount',
                            'price', 'cost', 'revenue', 'total', '合计', '小计',
                            '单价', '总价', '税额', '货款'],
        'date': ['日期', '时间', 'date', 'time', '创建时间', '更新时间', '生日',
                 '出生日期', '签约日期', '到期日'],
        'boolean': ['是否', '状态', 'status', 'active', 'enabled', 'flag',
                    '有效', '启用', '完成'],
        'gender': ['性别', 'gender', 'sex'],
        'url': ['网址', '链接', 'url', 'link', 'website', '官网'],
    }

    EMAIL_RE = re.compile(r'^[^@\s]+@[^@\s]+\.[^@\s]+$')
    PHONE_RE = re.compile(r'^[\d\s\-\+\(\)\.]{7,20}$')
    URL_RE = re.compile(r'^https?://')
    ID_CARD_RE = re.compile(r'^\d{17}[\dXx]$')

    @classmethod
    def detect(cls, header_name, sample_values):
        """推断列类型"""
        if not header_name:
            return 'general_text'

        h = str(header_name).strip().lower()

        # 1. 表头关键词匹配
        for col_type, keywords in cls.HEADER_KEYWORDS.items():
            for kw in keywords:
                if kw in h:
                    return col_type

        # 2. 数据采样分析
        samples = [str(v).strip() for v in sample_values
                   if v is not None and str(v).strip()][:50]
        if not samples:
            return 'general_text'

        email_count = sum(1 for s in samples if cls.EMAIL_RE.match(s))
        phone_count = sum(1 for s in samples if cls.PHONE_RE.match(s))
        url_count = sum(1 for s in samples if cls.URL_RE.match(s))
        id_count = sum(1 for s in samples if cls.ID_CARD_RE.match(s))

        total = len(samples)
        if email_count / total > 0.5:
            return 'email'
        if url_count / total > 0.5:
            return 'url'
        if id_count / total > 0.3:
            return 'id_number'
        if phone_count / total > 0.5:
            return 'phone'

        return 'general_text'

    @classmethod
    def detect_all_columns(cls, ws):
        """推断工作表所有列的类型"""
        col_types = {}
        if ws.max_row is None or ws.max_row < 1:
            return col_types

        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        for col_idx, header in enumerate(headers, 1):
            samples = []
            for row_idx in range(2, min(ws.max_row + 1, 52)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    samples.append(cell.value)

            col_letter = get_column_letter(col_idx)
            col_types[col_letter] = cls.detect(header, samples)

        return col_types


# ============================================================
#  处理函数 — 全部修正版
# ============================================================

# ── 一、字符编码 ──

def full_to_half_safe(text):
    """全角→半角，排除日文片假名(FF65-FF9F)和半角片假名"""
    result = []
    for ch in text:
        code = ord(ch)
        # 全角ASCII: FF01-FF5E → 半角 0021-007E
        if 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        elif code == 0x3000:  # 全角空格
            result.append(' ')
        # 不转换: 全角片假名(FF65-FF9F), 半角/全角形(FFA0-FFEF)
        else:
            result.append(ch)
    return ''.join(result)


def remove_zero_width_chars(text):
    zw = '\u200b\u200c\u200d\ufeff\u2060\u180e'
    for ch in zw:
        text = text.replace(ch, '')
    return text


def remove_control_chars(text):
    result = []
    for ch in text:
        c = ord(ch)
        if (0x00 <= c <= 0x08) or (0x0B <= c <= 0x0C) or (0x0E <= c <= 0x1F) or c == 0x7F:
            continue
        result.append(ch)
    return ''.join(result)


def remove_soft_hyphen(text):
    return text.replace('\u00ad', '')


def remove_bidi_marks(text):
    return text.replace('\u200e', '').replace('\u200f', '')


# ── 三、中文 ──

CHINESE_PUNCT_MAP = {
    '，': ',', '。': '.', '；': ';', '：': ':', '！': '!', '？': '?',
    '\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2019': "'",
}

# 安全的括号映射: 只转换中文圆括号变体, 绝不动 [] {} <> 《》
SAFE_BRACKET_MAP = {
    '（': '(', '）': ')',
    '﹙': '(', '﹚': ')',
}

CHINESE_NUM_MAP = {
    '零': 0, '一': 1, '二': 2, '三': 3, '四': 4,
    '五': 5, '六': 6, '七': 7, '八': 8, '九': 9,
    '十': 10, '百': 100, '千': 1000, '万': 10000, '亿': 100000000,
    '两': 2, '〇': 0,
}

CHINESE_FINANCIAL_MAP = {
    '壹': 1, '贰': 2, '叁': 3, '肆': 4, '伍': 5,
    '陆': 6, '柒': 7, '捌': 8, '玖': 9, '拾': 10,
    '佰': 100, '仟': 1000, '萬': 10000, '億': 100000000,
}


def _parse_chinese_number(s, num_map):
    total = 0
    section = 0
    current = 0
    for ch in s:
        val = num_map.get(ch)
        if val is None:
            return s
        if val >= 100000000:
            if section == 0 and current == 0:
                current = 1
            section += current
            total += section * val
            section = 0
            current = 0
        elif val >= 10000:
            if section == 0 and current == 0:
                current = 1
            section += current
            total += section * val
            section = 0
            current = 0
        elif val in (10, 100, 1000):
            if current == 0:
                current = 1
            section += current * val
            current = 0
        else:
            current = val
    total += section + current
    return total


def chinese_num_to_arabic(text):
    """中文数字→阿拉伯(只转换带单位的,如三百六十五)"""
    pattern = re.compile(r'[零一二三四五六七八九两〇]+(?:[十百千万亿][零一二三四五六七八九两〇]*)+')

    def _convert(m):
        s = m.group()
        try:
            result = _parse_chinese_number(s, CHINESE_NUM_MAP)
            return str(result) if isinstance(result, int) else s
        except Exception:
            return s

    pattern2 = re.compile(r'(?<![一-龥])[零一二三四五六七八九〇]{4,}(?![一-龥])')

    def _convert_digits(m):
        return ''.join(str(CHINESE_NUM_MAP.get(c, c)) for c in m.group())

    text = pattern.sub(_convert, text)
    text = pattern2.sub(_convert_digits, text)
    return text


def chinese_financial_to_arabic(text):
    combined = {**CHINESE_NUM_MAP, **CHINESE_FINANCIAL_MAP}
    pattern = re.compile(r'[壹贰叁肆伍陆柒捌玖零]+(?:[拾佰仟萬億][壹贰叁肆伍陆柒捌玖零]*)+')

    def _convert(m):
        try:
            result = _parse_chinese_number(m.group(), combined)
            return str(result) if isinstance(result, int) else m.group()
        except Exception:
            return m.group()

    return pattern.sub(_convert, text)


CHINESE_REGION_SUFFIXES = ['省', '市', '自治区', '自治州', '特别行政区', '地区', '盟']

def normalize_chinese_region(text):
    for suffix in CHINESE_REGION_SUFFIXES:
        if text.endswith(suffix) and len(text) > len(suffix) + 1:
            stripped = text[:-len(suffix)]
            if len(stripped) >= 2:
                return stripped
    return text


# ── 四、标点 ──

DASH_CHARS = '—–‐‑⁃―‒'
QUOTE_CHARS_DOUBLE = '\u201c\u201d\u300c\u300d\u300e\u300f\ufe41\ufe42\ufe43\ufe44'
QUOTE_CHARS_SINGLE = "\u2018\u2019\u201a\u201b"


def normalize_brackets_safe(text):
    """安全括号标准化: 只转换中文圆括号变体"""
    for old, new in SAFE_BRACKET_MAP.items():
        text = text.replace(old, new)
    return text


def normalize_quotes(text, style="double"):
    for ch in QUOTE_CHARS_DOUBLE:
        text = text.replace(ch, '"' if style == "double" else "'")
    for ch in QUOTE_CHARS_SINGLE:
        text = text.replace(ch, "'" if style == "single" else "'")
    return text


def normalize_dashes(text):
    for ch in DASH_CHARS:
        text = text.replace(ch, '-')
    return text


def normalize_ellipsis(text, style="three_dots"):
    if style == "three_dots":
        text = text.replace('\u2026', '...')
    else:
        text = re.sub(r'\.{3,}', '\u2026', text)
    return text


def remove_consecutive_punctuation(text):
    return re.sub(r'([!?.,;:])\1+', r'\1', text)


def remove_special_symbols(text):
    return re.sub(r'[™®©]', '', text)


# ── 五、数字 (安全版) ──

_SAFE_NUMBER_RE = re.compile(r'^[$¥€£]?\s*-?\d{1,3}(,\d{3})+(\.\d+)?$')


def remove_thousand_separators_safe(text):
    """安全千分位清除: 只在整个值看起来像格式化数字时才处理"""
    stripped = text.strip()
    if _SAFE_NUMBER_RE.match(stripped):
        return stripped.replace(',', '')
    return text


CURRENCY_SYMBOLS = ['$', '¥', '€', '£', '₩', '₹', '₽', '₺', '₴', '₦', '₫', '₿']
CURRENCY_WORDS = ['HK$', 'US$', 'A$', 'C$', 'S$', 'NZ$', 'RMB', 'CNY', 'USD', 'EUR', 'GBP', 'JPY']


def remove_currency_symbols(text):
    for w in CURRENCY_WORDS:
        text = text.replace(w, '')
    for s in CURRENCY_SYMBOLS:
        text = text.replace(s, '')
    return text.strip()


def scientific_to_number(text):
    pattern = re.compile(r'^[+-]?\d+\.?\d*[eE][+-]?\d+$')
    if pattern.match(text.strip()):
        try:
            num = float(text.strip())
            if abs(num) > 1e15:
                return text  # 大数不转换,避免精度丢失
            if num == int(num):
                return str(int(num))
            return f"{num:.10f}".rstrip('0').rstrip('.')
        except ValueError:
            pass
    return text


def accounting_negative(text):
    m = re.match(r'^\((\d[\d,.]*)\)$', text.strip())
    if m:
        return '-' + m.group(1)
    return text


def number_with_unit(text):
    unit_map = {'万': 10000, '亿': 100000000, '千': 1000, '百': 100}
    pattern = re.compile(r'(\d+\.?\d*)\s*([万亿千百])')

    def _replace(m):
        num = float(m.group(1))
        unit = unit_map.get(m.group(2), 1)
        result = num * unit
        return str(int(result)) if result == int(result) else str(result)

    return pattern.sub(_replace, text)


# ── 六、日期时间 ──

DATE_PATTERNS = [
    (re.compile(r'^(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})$'), 'YMD'),
    (re.compile(r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$'), 'DMY_or_MDY'),
    (re.compile(r'^(\d{4})年(\d{1,2})月(\d{1,2})日?$'), 'YMD'),
    (re.compile(r'^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+(\d{1,2}),?\s+(\d{4})$', re.I), 'MONTH_D_Y'),
    (re.compile(r'^(\d{1,2})\s+(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\.?\s+(\d{4})$', re.I), 'D_MONTH_Y'),
]

MONTH_MAP = {
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
}


def normalize_date_string(text, fmt="YYYY-MM-DD"):
    text = text.strip()
    y, m, d = None, None, None
    for pattern, ptype in DATE_PATTERNS:
        match = pattern.match(text)
        if not match:
            continue
        groups = match.groups()
        if ptype == 'YMD':
            y, m, d = int(groups[0]), int(groups[1]), int(groups[2])
        elif ptype == 'DMY_or_MDY':
            a, b, year = int(groups[0]), int(groups[1]), int(groups[2])
            if a > 12:
                d, m, y = a, b, year
            elif b > 12:
                m, d, y = a, b, year
            else:
                m, d, y = a, b, year
        elif ptype == 'MONTH_D_Y':
            m = MONTH_MAP.get(groups[0][:3].lower())
            d, y = int(groups[1]), int(groups[2])
        elif ptype == 'D_MONTH_Y':
            d = int(groups[0])
            m = MONTH_MAP.get(groups[1][:3].lower())
            y = int(groups[2])
        break
    if y is None or m is None or d is None:
        return text
    try:
        date(y, m, d)
    except ValueError:
        return text
    fmt_map = {
        "YYYY-MM-DD": f"{y:04d}-{m:02d}-{d:02d}",
        "DD/MM/YYYY": f"{d:02d}/{m:02d}/{y:04d}",
        "MM/DD/YYYY": f"{m:02d}/{d:02d}/{y:04d}",
        "YYYY年MM月DD日": f"{y:04d}年{m:02d}月{d:02d}日",
    }
    return fmt_map.get(fmt, text)


def normalize_time_string(text, fmt="HH:MM:SS"):
    text = text.strip()
    m12 = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM|am|pm|a\.m\.|p\.m\.)$', text)
    if m12:
        h, mi = int(m12.group(1)), int(m12.group(2))
        s = int(m12.group(3)) if m12.group(3) else 0
        ampm = m12.group(4).lower().replace('.', '')
        if ampm == 'pm' and h != 12:
            h += 12
        elif ampm == 'am' and h == 12:
            h = 0
    else:
        m24 = re.match(r'^(\d{1,2}):(\d{2})(?::(\d{2}))?$', text)
        if not m24:
            return text
        h, mi = int(m24.group(1)), int(m24.group(2))
        s = int(m24.group(3)) if m24.group(3) else 0
    if not (0 <= h <= 23 and 0 <= mi <= 59 and 0 <= s <= 59):
        return text
    if fmt == "HH:MM:SS":
        return f"{h:02d}:{mi:02d}:{s:02d}"
    elif fmt == "HH:MM":
        return f"{h:02d}:{mi:02d}"
    elif fmt == "12hour":
        ampm = "AM" if h < 12 else "PM"
        h12 = h % 12 or 12
        return f"{h12}:{mi:02d} {ampm}"
    return text


# ── 七、公司名称 (带上下文守卫) ──

COMPANY_SUFFIX_ABBR = OrderedDict([
    (r'\b(limited|ltd\.?)\b', 'ltd'),
    (r'\b(corporation|corp\.?)\b', 'corp'),
    (r'\b(company|co\.?)\b', 'co'),
    (r'\b(incorporated|inc\.?)\b', 'inc'),
    (r'\b(holdings?)\b', 'holding'),
    (r'\b(international|intl\.?)\b', 'intl'),
    (r'\b(enterprises?)\b', 'enterprise'),
    (r'\b(investments?)\b', 'investment'),
    (r'\b(technolog(?:y|ies)|tech)\b', 'tech'),
    (r'\b(management|mgmt\.?)\b', 'mgmt'),
    (r'\b(development|dev\.?)\b', 'dev'),
    (r'\b(private|pvt\.?)\b', 'pvt'),
    (r'\b(manufacturing|mfg\.?)\b', 'mfg'),
    (r'\b(laborator(?:y|ies)|labs?)\b', 'lab'),
    (r'\b(associates?|assoc\.?)\b', 'assoc'),
    (r'\b(industr(?:y|ies)|industrie)\b', 'industries'),
    (r'\b(solutions?)\b', 'solutions'),
    (r'\b(services?)\b', 'services'),
    (r'\b(systems?)\b', 'systems'),
])

COMPANY_SUFFIX_FULL = OrderedDict([
    (r'\b(limited|ltd\.?)\b', 'Limited'),
    (r'\b(corporation|corp\.?)\b', 'Corporation'),
    (r'\b(company|co\.?)\b', 'Company'),
    (r'\b(incorporated|inc\.?)\b', 'Incorporated'),
    (r'\b(holdings?)\b', 'Holdings'),
    (r'\b(international|intl\.?)\b', 'International'),
    (r'\b(enterprises?)\b', 'Enterprises'),
    (r'\b(investments?)\b', 'Investment'),
    (r'\b(technolog(?:y|ies)|tech)\b', 'Technology'),
    (r'\b(management|mgmt\.?)\b', 'Management'),
    (r'\b(development|dev\.?)\b', 'Development'),
    (r'\b(private|pvt\.?)\b', 'Private'),
    (r'\b(manufacturing|mfg\.?)\b', 'Manufacturing'),
    (r'\b(laborator(?:y|ies)|labs?)\b', 'Laboratories'),
    (r'\b(associates?|assoc\.?)\b', 'Associates'),
    (r'\b(industr(?:y|ies)|industrie)\b', 'Industries'),
])

COMPANY_INTL_SUFFIX = OrderedDict([
    (r'\bsdn\.?\s*bhd\.?\b', 'sdn bhd'),
    (r'\bpte\.?\s*ltd\.?\b', 'pte ltd'),
    (r'\bpty\.?\s*ltd\.?\b', 'pty ltd'),
    (r'\bl\.?l\.?c\.?\b', 'llc'),
    (r'\bl\.?l\.?p\.?\b', 'llp'),
    (r'\bp\.?l\.?c\.?\b', 'plc'),
    (r'\bg\.?m\.?b\.?h\.?\b', 'gmbh'),
    (r'\bs\.?a\.?r\.?l\.?\b', 'sarl'),
    (r'\bs\.?p\.?a\.?\b', 'spa'),
    (r'\bs\.?r\.?l\.?\b', 'srl'),
    (r'\bb\.?v\.?\b', 'bv'),
    (r'\bn\.?v\.?\b', 'nv'),
    (r'\bs\.?a\.?\b\.?(?!\w)', 'sa'),
    (r'\ba\.?g\.?\b\.?(?!\w)', 'ag'),
    (r'\bkgaa\b', 'kgaa'),
    (r'\bk\.?g\.?\b\.?(?!\w)', 'kg'),
    (r'\ba\.?b\.?\b\.?(?!\w)', 'ab'),
    (r'\ba/?s\b', 'as'),
    (r'\baps\b', 'aps'),
    (r'\bk\.?k\.?\b', 'kk'),
    (r'\b株式会社\b', 'kk'),
    (r'\bp\.?t\.?\b\.?(?!\w)', 'pt'),
    (r'\btbk\.?\b', 'tbk'),
    (r'\booo\b|ООО', 'ooo'),
    (r'\bo[üu]\.?\b', 'ou'),
])

CHINESE_COMPANY_MAP = [
    ('有限责任公司', '有限公司'),
    ('股份有限公司', '有限公司'),
]

COMPANY_STOP_WORDS = {'the', 'of', 'and', '&', 'a', 'an'}


def _has_company_suffix(text):
    """检查文本是否包含公司后缀关键词"""
    t = text.lower()
    keywords = ['ltd', 'limited', 'corp', 'corporation', 'inc', 'incorporated',
                'co.', 'company', 'llc', 'llp', 'plc', 'gmbh', 'sa', 'ag',
                'holding', 'enterprise', 'group', 'kk', '有限', '公司',
                '集团', '股份', '企业']
    return any(kw in t for kw in keywords)


def normalize_company_suffix_safe(text, style="abbreviation"):
    """公司后缀标准化 — 带上下文守卫"""
    if '@' in text or '://' in text:
        return text
    if len(text) < 3 or len(text) > 200:
        return text
    if not _has_company_suffix(text):
        return text
    mapping = COMPANY_SUFFIX_ABBR if style == "abbreviation" else COMPANY_SUFFIX_FULL
    for pattern, replacement in mapping.items():
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    return text


def normalize_company_intl_safe(text):
    if '@' in text or '://' in text:
        return text
    if not _has_company_suffix(text):
        return text
    for pattern, replacement in COMPANY_INTL_SUFFIX.items():
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    return text


def normalize_chinese_company(text):
    for old, new in CHINESE_COMPANY_MAP:
        text = text.replace(old, new)
    return text


def remove_company_stop_words_safe(text):
    """去除公司停用词 — 只在看起来像公司名的文本中"""
    if not _has_company_suffix(text):
        return text
    words = text.split()
    if len(words) <= 2:
        return text
    alpha_words = [w for w in words if re.match(r'^[a-zA-Z]+$', w)]
    if len(alpha_words) < 2:
        return text
    filtered = [w for w in words if w.lower() not in COMPANY_STOP_WORDS]
    return ' '.join(filtered) if filtered else text


def normalize_chinese_city_brackets(text):
    return re.sub(r'（([^）]+)）', r'(\1)', text)


# ── 八、通用文本 ──

SPELLING_US_TO_UK = {
    'color': 'colour', 'center': 'centre', 'organization': 'organisation',
    'realize': 'realise', 'analyze': 'analyse', 'defense': 'defence',
    'license': 'licence', 'catalog': 'catalogue', 'dialog': 'dialogue',
    'favor': 'favour', 'honor': 'honour', 'labor': 'labour',
    'neighbor': 'neighbour', 'behavior': 'behaviour', 'humor': 'humour',
    'fiber': 'fibre', 'liter': 'litre', 'meter': 'metre', 'theater': 'theatre',
}
SPELLING_UK_TO_US = {v: k for k, v in SPELLING_US_TO_UK.items()}

ADDRESS_EXPAND = {
    'st': 'street', 'st.': 'street', 'ave': 'avenue', 'ave.': 'avenue',
    'blvd': 'boulevard', 'blvd.': 'boulevard', 'rd': 'road', 'rd.': 'road',
    'dr': 'drive', 'dr.': 'drive', 'ln': 'lane', 'ln.': 'lane',
    'ct': 'court', 'ct.': 'court', 'pl': 'place', 'pl.': 'place',
    'pkwy': 'parkway', 'hwy': 'highway',
}
ADDRESS_ABBREV = {v: k.rstrip('.') for k, v in ADDRESS_EXPAND.items() if '.' not in k}

# 修正版Emoji正则: 不包含Box Drawing(2500-257F)和数学符号
EMOJI_PATTERN_SAFE = re.compile(
    "["
    "\U0001F600-\U0001F64F"  # Emoticons
    "\U0001F300-\U0001F5FF"  # Symbols & Pictographs
    "\U0001F680-\U0001F6FF"  # Transport & Map
    "\U0001F1E0-\U0001F1FF"  # Flags
    "\U0001F900-\U0001F9FF"  # Supplemental Symbols
    "\U0001FA00-\U0001FA6F"  # Chess Symbols
    "\U0001FA70-\U0001FAFF"  # Symbols Extended-A
    "\U00002702-\U000027B0"  # Dingbats
    "\U0001f926-\U0001f937"
    "]+",
    flags=re.UNICODE,
)

# 严格空值集: 只匹配明确的空值表达
NULL_VALUES_STRICT = {'n/a', 'na', 'null', 'none', '#n/a', '#na', 'nil', 'nan'}
# 扩展空值集(需用户显式启用)
NULL_VALUES_EXTENDED = {'n/a', 'na', 'null', 'none', '#n/a', '#na', 'nil', 'nan',
                        '-', '—', '无', '暂无', '空', '未填写', '未知', '不详',
                        '/', '--', '---'}

BOOLEAN_TRUE = {'y', 'yes', '是', 'true', '1', '✓', '√', 'on', 'si'}
BOOLEAN_FALSE = {'n', 'no', '否', 'false', '0', '✗', '×', 'off'}

GENDER_MALE = {'m', 'male', '男', '先生', 'mr', 'mr.'}
GENDER_FEMALE = {'f', 'female', '女', '女士', '小姐', 'ms', 'ms.', 'mrs', 'mrs.', 'miss'}


def normalize_spelling(text, style="american"):
    mapping = SPELLING_UK_TO_US if style == "american" else SPELLING_US_TO_UK
    words = text.split()
    result = []
    for w in words:
        lw = w.lower()
        if lw in mapping:
            nw = mapping[lw]
            if w[0].isupper():
                nw = nw.capitalize()
            result.append(nw)
        else:
            result.append(w)
    return ' '.join(result)


def remove_emoji_safe(text):
    return EMOJI_PATTERN_SAFE.sub('', text)


def normalize_html_entities(text):
    return html_module.unescape(text)


def normalize_url(text, style="remove_protocol"):
    if style == "remove_protocol":
        return re.sub(r'https?://(www\.)?', '', text)
    elif style == "extract_domain":
        m = re.match(r'https?://(?:www\.)?([^/]+)', text)
        return m.group(1) if m else text
    return text


def normalize_email(text):
    if re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', text.strip()):
        return text.strip().lower()
    return text


def normalize_boolean(text, style="yes_no"):
    lower = text.strip().lower()
    if lower in BOOLEAN_TRUE:
        return {"yes_no": "Yes", "true_false": "True", "1_0": "1", "是_否": "是"}[style]
    elif lower in BOOLEAN_FALSE:
        return {"yes_no": "No", "true_false": "False", "1_0": "0", "是_否": "否"}[style]
    return text


def normalize_null_strict(text, style="empty"):
    """严格空值标准化: 只匹配明确的空值"""
    if text.strip().lower() in NULL_VALUES_STRICT:
        return "" if style == "empty" else style
    return text


def normalize_gender(text, style="M_F"):
    lower = text.strip().lower()
    if lower in GENDER_MALE:
        return {"M_F": "M", "Male_Female": "Male", "男_女": "男"}[style]
    elif lower in GENDER_FEMALE:
        return {"M_F": "F", "Male_Female": "Female", "男_女": "女"}[style]
    return text


def normalize_address(text, style="expand"):
    mapping = ADDRESS_EXPAND if style == "expand" else ADDRESS_ABBREV
    words = text.split()
    result = []
    for w in words:
        lw = w.lower()
        if lw in mapping:
            nw = mapping[lw]
            if w[0].isupper():
                nw = nw.capitalize()
            result.append(nw)
        else:
            result.append(w)
    return ' '.join(result)


# ── 九、电话与证件 ──

def normalize_phone(text):
    if re.search(r'\d', text) and len(re.sub(r'\D', '', text)) >= 7:
        cleaned = re.sub(r'[\s\-\(\)\.\+]', '', text)
        if cleaned.isdigit() and len(cleaned) >= 7:
            return cleaned
    return text


def normalize_id_card(text):
    if re.match(r'^\d{17}[xX]$', text.strip()):
        return text.strip()[:-1] + 'X'
    return text


def normalize_credit_code(text):
    if re.match(r'^[0-9A-Za-z]{18}$', text.strip()):
        return text.strip().upper()
    return text


def normalize_bank_card(text):
    cleaned = re.sub(r'[\s\-]', '', text)
    if cleaned.isdigit() and 13 <= len(cleaned) <= 19:
        return cleaned
    return text


# ============================================================
#  核心处理引擎 — 两阶段模型
# ============================================================

class ChangeLogger:
    def __init__(self):
        self.logs = []
        self.warnings = []
        self.skipped = []

    def log(self, sheet, cell_ref, original, processed, rule):
        if str(original) != str(processed):
            self.logs.append({
                'sheet': sheet, 'cell': cell_ref,
                'original': str(original), 'processed': str(processed),
                'rule': rule, 'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })

    def warn(self, sheet, cell_ref, value, msg):
        self.warnings.append({
            'sheet': sheet, 'cell': cell_ref, 'value': str(value), 'message': msg
        })

    def skip(self, sheet, cell_ref, value, reason):
        self.skipped.append({
            'sheet': sheet, 'cell': cell_ref, 'value': str(value)[:50], 'reason': reason
        })

    def get_summary(self):
        summary = {}
        for log in self.logs:
            rule = log['rule']
            summary[rule] = summary.get(rule, 0) + 1
        return summary

    def export_to_workbook(self, filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = "变更日志"
        headers = ['Sheet名', '单元格', '原始值', '处理后', '处理规则', '时间']
        for col, header in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=header)
            c.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            c.fill = openpyxl.styles.PatternFill('solid', fgColor='4472C4')
            c.alignment = openpyxl.styles.Alignment(horizontal='center')
        for i, log in enumerate(self.logs, 2):
            ws.cell(row=i, column=1, value=log['sheet'])
            ws.cell(row=i, column=2, value=log['cell'])
            ws.cell(row=i, column=3, value=log['original'])
            ws.cell(row=i, column=4, value=log['processed'])
            ws.cell(row=i, column=5, value=log['rule'])
            ws.cell(row=i, column=6, value=log['timestamp'])

        # 汇总
        ws2 = wb.create_sheet("汇总统计")
        ws2.cell(row=1, column=1, value="处理规则").font = openpyxl.styles.Font(bold=True)
        ws2.cell(row=1, column=2, value="变更次数").font = openpyxl.styles.Font(bold=True)
        for i, (rule, count) in enumerate(self.get_summary().items(), 2):
            ws2.cell(row=i, column=1, value=rule)
            ws2.cell(row=i, column=2, value=count)

        # 警告
        if self.warnings:
            ws3 = wb.create_sheet("⚠️ 警告")
            ws3.append(["Sheet", "单元格", "值", "警告信息"])
            for w in self.warnings:
                ws3.append([w['sheet'], w['cell'], w['value'], w['message']])

        # 跳过记录
        if self.skipped:
            ws4 = wb.create_sheet("跳过记录")
            ws4.append(["Sheet", "单元格", "值(截断)", "跳过原因"])
            for s in self.skipped:
                ws4.append([s['sheet'], s['cell'], s['value'], s['reason']])

        # 列宽
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                cl = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                sheet.column_dimensions[cl].width = min(max_len + 4, 60)
            sheet.freeze_panes = 'A2'

        wb.save(filepath)
        return len(self.logs)


def is_formula(value):
    """检测公式"""
    return isinstance(value, str) and value.startswith('=')


def should_apply_rule(rule_id, col_type, settings):
    """检查规则是否应对该列类型应用"""
    if rule_id in RULE_COLUMN_APPLICABILITY:
        applicable_types = RULE_COLUMN_APPLICABILITY[rule_id]
        if col_type not in applicable_types:
            return False
    return settings.get(rule_id, {}).get("enabled", False)


def process_text_value(text, settings, logger, sheet_name, cell_ref, col_type='general_text'):
    """对单个文本值应用所有启用的(且适用于该列类型的)规则"""
    if not isinstance(text, str) or not text:
        return text

    def _apply(rule_id, func, desc, *args):
        nonlocal text
        if not should_apply_rule(rule_id, col_type, settings):
            return
        new = func(text, *args) if args else func(text)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, desc)
            text = new

    # ── 一、字符编码 ──
    _apply("1.09_control_chars", remove_control_chars, "1.09 控制字符清除")
    _apply("1.08_zero_width_chars", remove_zero_width_chars, "1.08 零宽字符清除")
    _apply("1.11_soft_hyphen", remove_soft_hyphen, "1.11 软连字符清除")
    _apply("1.12_bidi_marks", remove_bidi_marks, "1.12 双向标记清除")

    if should_apply_rule("1.10_unicode_normalize", col_type, settings):
        form = settings["1.10_unicode_normalize"].get("option", "NFC")
        new = unicodedata.normalize(form, text)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, f"1.10 Unicode归一化({form})")
            text = new

    _apply("1.01_fullwidth_to_halfwidth", full_to_half_safe, "1.01 全角→半角(安全)")

    if should_apply_rule("1.02_fullwidth_space", col_type, settings):
        new = text.replace('\u3000', ' ')
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "1.02 全角空格→半角")
            text = new

    if should_apply_rule("1.03_nbsp_to_space", col_type, settings):
        new = text.replace('\xa0', ' ')
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "1.03 NBSP→空格")
            text = new

    if should_apply_rule("1.04_tab_to_space", col_type, settings):
        new = text.replace('\t', ' ')
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "1.04 Tab→空格")
            text = new

    if should_apply_rule("1.05_newline_to_space", col_type, settings):
        new = text.replace('\r\n', ' ').replace('\r', ' ').replace('\n', ' ')
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "1.05 换行→空格")
            text = new

    if should_apply_rule("1.06_collapse_spaces", col_type, settings):
        new = re.sub(r' {2,}', ' ', text)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "1.06 连续空格合并")
            text = new

    # ── HTML实体(标点处理前) ──
    _apply("8.07_html_entities", normalize_html_entities, "8.07 HTML实体转换")

    # ── 三、中文 ──
    if should_apply_rule("3.01_traditional_to_simplified", col_type, settings) and HAS_OPENCC:
        try:
            cc = OpenCC('t2s')
            new = cc.convert(text)
            if new != text:
                logger.log(sheet_name, cell_ref, text, new, "3.01 繁体→简体")
                text = new
        except Exception:
            pass

    if should_apply_rule("3.05_simplified_to_traditional", col_type, settings) and HAS_OPENCC:
        try:
            cc = OpenCC('s2t')
            new = cc.convert(text)
            if new != text:
                logger.log(sheet_name, cell_ref, text, new, "3.05 简体→繁体")
                text = new
        except Exception:
            pass

    _apply("3.06_chinese_number_to_arabic", chinese_num_to_arabic, "3.06 中文数字→阿拉伯")
    _apply("3.07_chinese_financial_number", chinese_financial_to_arabic, "3.07 大写金额→阿拉伯")

    if should_apply_rule("3.02_chinese_punctuation", col_type, settings):
        new = text
        for old, repl in CHINESE_PUNCT_MAP.items():
            new = new.replace(old, repl)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "3.02 中文标点→英文")
            text = new

    if should_apply_rule("3.03_chinese_brackets", col_type, settings):
        new = normalize_brackets_safe(text)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "3.03 中文圆括号→英文")
            text = new

    if should_apply_rule("3.04_chinese_comma_dot", col_type, settings):
        new = text.replace('、', ',')
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "3.04 顿号→逗号")
            text = new

    _apply("3.08_chinese_region_normalize", normalize_chinese_region, "3.08 省市区名标准化")

    # ── 四、标点 ──
    if should_apply_rule("4.01_brackets_normalize", col_type, settings):
        new = normalize_brackets_safe(text)  # 使用安全版: 只转中文圆括号
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "4.01 括号统一(安全)")
            text = new

    if should_apply_rule("4.02_tilde_normalize", col_type, settings):
        new = text.replace('～', '~')
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "4.02 波浪号统一")
            text = new

    if should_apply_rule("4.03_quotes_normalize", col_type, settings):
        style = settings["4.03_quotes_normalize"].get("option", "double")
        new = normalize_quotes(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "4.03 引号统一")
            text = new

    _apply("4.04_dash_normalize", normalize_dashes, "4.04 破折号统一")

    if should_apply_rule("4.05_ellipsis_normalize", col_type, settings):
        style = settings["4.05_ellipsis_normalize"].get("option", "three_dots")
        new = normalize_ellipsis(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "4.05 省略号统一")
            text = new

    _apply("4.06_consecutive_punctuation", remove_consecutive_punctuation, "4.06 连续标点去重")
    _apply("4.07_special_symbols_clean", remove_special_symbols, "4.07 特殊符号清理")

    # ── 五、数字 ──
    _apply("5.06_negative_accounting", accounting_negative, "5.06 会计负数转换")
    _apply("5.02_thousand_separator", remove_thousand_separators_safe, "5.02 千分位清除(安全)")

    if should_apply_rule("5.03_currency_symbols", col_type, settings):
        opt = settings["5.03_currency_symbols"].get("option", "remove")
        if opt == "remove":
            new = remove_currency_symbols(text)
            if new != text:
                logger.log(sheet_name, cell_ref, text, new, "5.03 货币符号清除")
                text = new

    _apply("5.05_scientific_notation", scientific_to_number, "5.05 科学计数法还原")
    _apply("5.07_number_with_unit", number_with_unit, "5.07 带单位数字拆分")

    if should_apply_rule("5.04_percent_handling", col_type, settings):
        opt = settings["5.04_percent_handling"].get("option", "keep_string")
        if opt == "to_decimal":
            m = re.match(r'^(\d+\.?\d*)\s*%$', text.strip())
            if m:
                new = str(float(m.group(1)) / 100)
                logger.log(sheet_name, cell_ref, text, new, "5.04 百分号转小数")
                text = new

    # ── 六、日期时间 ──
    if should_apply_rule("6.01_date_normalize", col_type, settings):
        fmt = settings["6.01_date_normalize"].get("option", "YYYY-MM-DD")
        new = normalize_date_string(text, fmt)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, f"6.01 日期统一({fmt})")
            text = new

    if should_apply_rule("6.02_time_normalize", col_type, settings):
        fmt = settings["6.02_time_normalize"].get("option", "HH:MM:SS")
        new = normalize_time_string(text, fmt)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, f"6.02 时间统一({fmt})")
            text = new

    # ── 七、公司名称(带守卫) ──
    _apply("7.05_chinese_city_brackets", normalize_chinese_city_brackets, "7.05 中文括号地名")
    _apply("7.03_chinese_company_type", normalize_chinese_company, "7.03 中文公司类型")

    if should_apply_rule("7.01_company_suffix_normalize", col_type, settings):
        style = settings["7.01_company_suffix_normalize"].get("option", "abbreviation")
        new = normalize_company_suffix_safe(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "7.01 公司后缀统一(安全)")
            text = new

    _apply("7.02_company_intl_suffix", normalize_company_intl_safe, "7.02 国际公司后缀(安全)")
    _apply("7.04_company_stop_words", remove_company_stop_words_safe, "7.04 停用词去除(安全)")

    # ── 八、通用文本 ──
    _apply("8.06_emoji_remove", remove_emoji_safe, "8.06 Emoji清除(安全)")
    _apply("8.09_email_normalize", normalize_email, "8.09 邮箱标准化")

    if should_apply_rule("8.03_spelling_variant", col_type, settings):
        style = settings["8.03_spelling_variant"].get("option", "american")
        new = normalize_spelling(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "8.03 拼写变体统一")
            text = new

    if should_apply_rule("8.05_diacritics_remove", col_type, settings) and HAS_UNIDECODE:
        # 安全版变音符号清除: 只对非CJK字符使用unidecode
        # unidecode 会把中文字符转成拼音(如 "中国"→"Zhong Guo"), 这是灾难性的!
        # 所以必须逐字符处理, 跳过CJK/日文/韩文字符
        new_chars = []
        for ch in text:
            cp = ord(ch)
            # 跳过CJK统一表意文字及相关区域
            if (0x2E80 <= cp <= 0x9FFF or   # CJK部首/表意文字
                0xF900 <= cp <= 0xFAFF or   # CJK兼容表意
                0xFE30 <= cp <= 0xFE4F or   # CJK兼容形式
                0x20000 <= cp <= 0x2FA1F or  # CJK扩展B-F
                0x3040 <= cp <= 0x30FF or   # 日文平假名/片假名
                0x31F0 <= cp <= 0x31FF or   # 片假名扩展
                0xAC00 <= cp <= 0xD7AF or   # 韩文音节
                0x1100 <= cp <= 0x11FF or   # 韩文字母
                0x3130 <= cp <= 0x318F or   # 韩文兼容字母
                0xFF00 <= cp <= 0xFFEF):    # 半角/全角形式
                new_chars.append(ch)
            else:
                new_chars.append(unidecode(ch))
        new = ''.join(new_chars)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "8.05 变音符号清除(保留CJK)")
            text = new

    if should_apply_rule("8.08_url_normalize", col_type, settings):
        if re.match(r'https?://', text):
            style = settings["8.08_url_normalize"].get("option", "remove_protocol")
            new = normalize_url(text, style)
            if new != text:
                logger.log(sheet_name, cell_ref, text, new, "8.08 URL清理")
                text = new

    if should_apply_rule("8.10_boolean_normalize", col_type, settings):
        style = settings["8.10_boolean_normalize"].get("option", "yes_no")
        new = normalize_boolean(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "8.10 布尔值统一")
            text = new

    if should_apply_rule("8.11_null_normalize", col_type, settings):
        style = settings["8.11_null_normalize"].get("option", "empty")
        new = normalize_null_strict(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "8.11 空值统一(严格)")
            text = new

    if should_apply_rule("8.12_gender_normalize", col_type, settings):
        style = settings["8.12_gender_normalize"].get("option", "M_F")
        new = normalize_gender(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "8.12 性别统一")
            text = new

    # ── 九、证件号 ──
    _apply("9.01_phone_normalize", normalize_phone, "9.01 电话标准化")
    _apply("9.02_id_card_normalize", normalize_id_card, "9.02 身份证X大写")
    _apply("9.03_credit_code_normalize", normalize_credit_code, "9.03 社会信用代码大写")
    _apply("9.04_bank_card_normalize", normalize_bank_card, "9.04 银行卡号清理")

    # ── 十二、地址 ──
    if should_apply_rule("12.01_address_abbreviation", col_type, settings):
        style = settings["12.01_address_abbreviation"].get("option", "expand")
        new = normalize_address(text, style)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "12.01 地址缩写统一")
            text = new

    # ── 二、大小写(最后执行) ──
    if should_apply_rule("2.01_case_transform", col_type, settings):
        opt = settings["2.01_case_transform"].get("option", "none")
        if opt == "lower":
            new = text.lower()
        elif opt == "upper":
            new = text.upper()
        elif opt == "title":
            new = text.title()
        else:
            new = text
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, f"2.01 大小写({opt})")
            text = new

    # ── 最终清理 ──
    if should_apply_rule("8.02_collapse_cell_spaces", col_type, settings):
        new = re.sub(r' {2,}', ' ', text)
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "8.02 空格合并")
            text = new

    if (should_apply_rule("8.01_strip_cell", col_type, settings) or
            should_apply_rule("1.07_strip_whitespace", col_type, settings)):
        new = text.strip()
        if new != text:
            logger.log(sheet_name, cell_ref, text, new, "1.07/8.01 首尾空白去除")
            text = new

    return text


# ============================================================
#  预检机制
# ============================================================

def pre_check(wb):
    """文件预检: 发现潜在风险并警告"""
    warnings = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        formula_count = 0
        long_num_count = 0
        merged_count = len(list(ws.merged_cells.ranges))

        for row in ws.iter_rows():
            for cell in row:
                if is_formula(cell.value):
                    formula_count += 1
                if isinstance(cell.value, float) and abs(cell.value) > 1e15:
                    long_num_count += 1

        if formula_count > 0:
            warnings.append(f"[{sheet_name}] 含 {formula_count} 个公式单元格(将自动跳过)")
        if long_num_count > 0:
            warnings.append(f"[{sheet_name}] 含 {long_num_count} 个大数值(可能是被Excel截断的长编号)")
        if merged_count > 0:
            warnings.append(f"[{sheet_name}] 含 {merged_count} 个合并区域(只处理主单元格)")
    return warnings


# ============================================================
#  两阶段处理核心
# ============================================================

def get_merged_non_master(ws):
    non_master = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    non_master.add((r, c))
    return non_master


def compute_proposals(ws, sheet_name, settings, col_types, logger):
    """阶段1: 扫描计算所有变更提案(只读)"""
    proposals = []
    non_master = get_merged_non_master(ws)

    for row in ws.iter_rows():
        for cell in row:
            if (cell.row, cell.column) in non_master:
                continue
            if cell.value is None:
                continue

            cell_ref = f"{get_column_letter(cell.column)}{cell.row}"
            col_letter = get_column_letter(cell.column)
            col_type = col_types.get(col_letter, 'general_text')

            # P0: 公式绝不触碰
            if is_formula(cell.value):
                logger.skip(sheet_name, cell_ref, cell.value, "公式单元格")
                continue

            # 数字类型: 保持原样
            if isinstance(cell.value, (int, float)):
                if isinstance(cell.value, float) and abs(cell.value) > 1e15:
                    logger.warn(sheet_name, cell_ref, cell.value,
                                "大数值,可能是被Excel截断的长编号(如身份证/卡号)")
                continue

            # 日期时间/布尔: 保持原样
            if isinstance(cell.value, (datetime, date, dt_time, bool)):
                continue

            # 字符串: 计算变更
            val = str(cell.value)
            temp_logger = ChangeLogger()
            new_val = process_text_value(val, settings, temp_logger, sheet_name, cell_ref, col_type)

            if new_val != val:
                # 收集每条规则产生的变更
                for log_entry in temp_logger.logs:
                    proposals.append(ChangeProposal(
                        sheet=sheet_name, cell_ref=cell_ref,
                        row=cell.row, col=cell.column,
                        original=val, proposed=new_val,
                        rule_id=log_entry['rule'], rule_desc=log_entry['rule']
                    ))
                # 汇总为一条最终变更(用于实际应用)
                # 但在审核界面展示每步规则
                # 这里我们用最终结果做一条总提案
                proposals.append(ChangeProposal(
                    sheet=sheet_name, cell_ref=cell_ref,
                    row=cell.row, col=cell.column,
                    original=val, proposed=new_val,
                    rule_id="_FINAL_", rule_desc="最终结果"
                ))

    # 只保留_FINAL_提案用于应用, 其余用于审核展示
    return proposals


def apply_confirmed_changes(wb, proposals):
    """阶段3: 只应用被确认的变更"""
    applied = 0
    for p in proposals:
        if p.rule_id != "_FINAL_":
            continue
        if not p.accepted and p.user_override is None:
            continue
        ws = wb[p.sheet]
        cell = ws.cell(row=p.row, column=p.col)

        # 保存原始格式
        orig_font = copy.copy(cell.font)
        orig_fill = copy.copy(cell.fill)
        orig_border = copy.copy(cell.border)
        orig_align = copy.copy(cell.alignment)
        orig_nf = cell.number_format
        orig_prot = copy.copy(cell.protection)

        # 写入新值
        cell.value = p.final_value

        # 恢复原始格式
        cell.font = orig_font
        cell.fill = orig_fill
        cell.border = orig_border
        cell.alignment = orig_align
        cell.number_format = orig_nf
        cell.protection = orig_prot

        applied += 1
    return applied


# ============================================================
#  审核界面
# ============================================================

class ReviewInterface:
    def __init__(self, proposals):
        # 只看最终结果提案
        self.all_proposals = proposals
        self.final_proposals = [p for p in proposals if p.rule_id == "_FINAL_"]
        self.detail_proposals = [p for p in proposals if p.rule_id != "_FINAL_"]

    def show_summary(self):
        print(f"\n  📊 共发现 {len(self.final_proposals)} 个单元格待变更")
        # 按sheet统计
        by_sheet = {}
        for p in self.final_proposals:
            by_sheet[p.sheet] = by_sheet.get(p.sheet, 0) + 1
        for sheet, count in by_sheet.items():
            print(f"     [{sheet}] {count} 处")

        # 按规则统计(用detail)
        by_rule = {}
        for p in self.detail_proposals:
            by_rule[p.rule_id] = by_rule.get(p.rule_id, 0) + 1
        if by_rule:
            print(f"\n  📋 规则触发统计:")
            for rule, count in sorted(by_rule.items()):
                print(f"     {rule}: {count} 次")

    def review_all(self):
        print("\n  审核模式:")
        print("  1. 逐条审核")
        print("  2. 查看全部后批量确认")
        print("  3. 全部接受")
        print("  4. 全部拒绝")
        choice = input("  > ").strip()

        if choice == '1':
            self._review_one_by_one()
        elif choice == '2':
            self._review_batch()
        elif choice == '3':
            pass  # 默认accepted=True
        elif choice == '4':
            for p in self.final_proposals:
                p.accepted = False

    def _review_one_by_one(self):
        for i, p in enumerate(self.final_proposals):
            print(f"\n  [{i+1}/{len(self.final_proposals)}]")
            print(f"  位置: {p.sheet} → {p.cell_ref}")
            print(f"  原始: [{p.original}]")
            print(f"  修改: [{p.proposed}]")

            # 显示涉及的规则
            rules = [d.rule_id for d in self.detail_proposals
                     if d.sheet == p.sheet and d.cell_ref == p.cell_ref]
            if rules:
                print(f"  规则: {', '.join(rules)}")

            action = input("  (a)接受 (r)拒绝 (e)自定义 (q)退出: ").strip().lower()
            if action == 'r':
                p.accepted = False
            elif action == 'e':
                custom = input("  输入自定义值: ")
                p.user_override = custom
            elif action == 'q':
                break

    def _review_batch(self):
        print(f"\n  全部变更预览 (共{len(self.final_proposals)}条):")
        for i, p in enumerate(self.final_proposals[:50]):
            print(f"  {i+1}. [{p.sheet}]{p.cell_ref}: [{p.original}] → [{p.proposed}]")
        if len(self.final_proposals) > 50:
            print(f"  ... 还有 {len(self.final_proposals)-50} 条")

        action = input("\n  (a)全部接受 (r)全部拒绝 (s)选择性拒绝(输入编号): ").strip().lower()
        if action == 'r':
            for p in self.final_proposals:
                p.accepted = False
        elif action.startswith('s'):
            reject_ids = input("  输入要拒绝的编号(逗号分隔): ").strip()
            for idx_str in reject_ids.split(','):
                try:
                    idx = int(idx_str.strip()) - 1
                    if 0 <= idx < len(self.final_proposals):
                        self.final_proposals[idx].accepted = False
                except ValueError:
                    pass


def export_review_report(proposals, filepath):
    """导出可审核的变更报告Excel"""
    final = [p for p in proposals if p.rule_id == "_FINAL_"]
    detail = [p for p in proposals if p.rule_id != "_FINAL_"]

    wb = Workbook()
    ws = wb.active
    ws.title = "变更审核"

    headers = ['序号', 'Sheet', '单元格', '涉及规则', '原始值', '建议修改为',
               '是否接受(Y/N)', '自定义修改(可选)']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='4472C4')

    for i, p in enumerate(final, 1):
        rules = [d.rule_id for d in detail
                 if d.sheet == p.sheet and d.cell_ref == p.cell_ref]
        ws.append([i, p.sheet, p.cell_ref, '; '.join(rules), p.original, p.proposed, 'Y', ''])

    # 数据验证
    dv = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f'G2:G{len(final)+1}')

    # 列宽
    for col in ws.columns:
        cl = get_column_letter(col[0].column)
        max_len = max((len(str(c.value or '')) for c in col), default=10)
        ws.column_dimensions[cl].width = min(max_len + 4, 60)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"  📋 审核报告已导出: {filepath}")


def load_review_decisions(review_filepath, proposals):
    """从审核报告加载用户决策"""
    final = [p for p in proposals if p.rule_id == "_FINAL_"]
    wb = load_workbook(review_filepath)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=False):
        idx = row[0].value
        if idx is None:
            continue
        idx = int(idx) - 1
        if idx >= len(final):
            continue
        accept = str(row[6].value or 'Y').strip().upper()
        custom = row[7].value
        final[idx].accepted = (accept == 'Y')
        if custom:
            final[idx].user_override = str(custom)
    return proposals


# ============================================================
#  格式美化 (可选, 全部默认关闭)
# ============================================================

def apply_output_formatting(ws, settings):
    """可选的格式美化, 全部默认关闭"""
    if not ws.max_row:
        return

    if settings.get("10.01_freeze_header", {}).get("enabled"):
        ws.freeze_panes = 'A2'

    if settings.get("10.02_header_style", {}).get("enabled"):
        hf = openpyxl.styles.Font(bold=True, color='FFFFFF', size=11)
        hfill = openpyxl.styles.PatternFill('solid', fgColor='4472C4')
        ha = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            if cell.value is not None:
                cell.font = hf
                cell.fill = hfill
                cell.alignment = ha

    if settings.get("10.03_auto_column_width", {}).get("enabled"):
        for col in ws.columns:
            max_len = 0
            cl = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    vs = str(cell.value)
                    char_len = sum(2 if ord(c) > 127 else 1 for c in vs)
                    max_len = max(max_len, char_len)
            ws.column_dimensions[cl].width = min(max_len + 4, 80)

    if settings.get("10.04_auto_filter", {}).get("enabled") and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions

    if settings.get("10.06_font_unify", {}).get("enabled"):
        font_str = settings["10.06_font_unify"].get("option", "Arial 11")
        parts = font_str.rsplit(' ', 1)
        fname = parts[0] if len(parts) == 2 else "Arial"
        fsize = int(parts[1]) if len(parts) == 2 else 11
        font = openpyxl.styles.Font(name=fname, size=fsize)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = font


def process_header(ws, settings, logger, sheet_name):
    if not settings.get("11.01_header_clean", {}).get("enabled"):
        return
    for cell in ws[1]:
        if cell.value and isinstance(cell.value, str):
            old = cell.value
            new = old.strip().replace('\n', ' ').replace('\r', ' ')
            new = re.sub(r'\s+', ' ', new)
            if new != old:
                cell_ref = f"{get_column_letter(cell.column)}1"
                logger.log(sheet_name, cell_ref, old, new, "11.01 表头清理")
                cell.value = new


# ============================================================
#  主处理流程 v3
# ============================================================

def process_file_v3(filepath, settings, auto_accept=False):
    """重构后的主处理流程: 预检→列检测→试算→审核→应用"""

    # ── 阶段0: 预检 ──
    print("\n  📋 阶段0: 文件预检...")
    wb = load_workbook(filepath, data_only=False, keep_links=True)

    warnings = pre_check(wb)
    if warnings:
        print("  ⚠️  预检发现以下问题:")
        for w in warnings:
            print(f"     {w}")
        if not auto_accept:
            if input("  是否继续? (Y/n): ").strip().lower() == 'n':
                return None

    # 选择sheet
    if auto_accept:
        sheets = wb.sheetnames
    else:
        sheets = select_sheets(wb)
    if not sheets:
        print("  ❌ 未选择任何工作表")
        return None

    logger = ChangeLogger()

    # ── 阶段1: 列类型检测 ──
    print("\n  🔍 阶段1: 分析列数据类型...")
    all_col_types = {}
    for sn in sheets:
        ws = wb[sn]
        col_types = ColumnTypeDetector.detect_all_columns(ws)
        all_col_types[sn] = col_types
        if col_types:
            for col, ctype in col_types.items():
                if ctype != 'general_text':
                    print(f"     [{sn}] 列{col}: {ctype}")

    if not auto_accept:
        print("\n  ℹ️  列类型会影响规则的应用范围(如公司名规则只对company_name列)")
        if input("  是否需要手动修正列类型? (y/N): ").strip().lower() == 'y':
            for sn in sheets:
                for col in all_col_types.get(sn, {}):
                    current = all_col_types[sn][col]
                    new_type = input(f"     [{sn}] 列{col} 当前={current}, 新类型(回车跳过): ").strip()
                    if new_type:
                        all_col_types[sn][col] = new_type

    # ── 阶段2: 试算变更 ──
    print("\n  ⚙️  阶段2: 计算变更提案...")
    all_proposals = []
    for sn in sheets:
        ws = wb[sn]
        process_header(ws, settings, logger, sn)
        col_types = all_col_types.get(sn, {})
        proposals = compute_proposals(ws, sn, settings, col_types, logger)
        all_proposals.extend(proposals)

    final_count = len([p for p in all_proposals if p.rule_id == "_FINAL_"])

    if final_count == 0:
        print("  ✅ 未发现需要变更的内容")
        if logger.warnings:
            print(f"  ⚠️  但有 {len(logger.warnings)} 条警告:")
            for w in logger.warnings[:10]:
                print(f"     [{w['sheet']}]{w['cell']}: {w['message']}")
        return filepath

    # ── 阶段3: 审核 ──
    print(f"\n  👁️  阶段3: 变更审核 (共{final_count}个单元格)")
    reviewer = ReviewInterface(all_proposals)
    reviewer.show_summary()

    if auto_accept:
        mode = '3'
    else:
        print("\n  选择审核方式:")
        print("  1. 终端交互审核")
        print("  2. 导出Excel审核报告")
        print("  3. 全部接受")
        mode = input("  > ").strip()

    if mode == '1':
        reviewer.review_all()
    elif mode == '2':
        review_path = filepath.replace('.xlsx', '_审核报告.xlsx').replace('.xlsm', '_审核报告.xlsx')
        export_review_report(all_proposals, review_path)
        input(f"  请编辑 {review_path} 后按回车继续...")
        all_proposals = load_review_decisions(review_path, all_proposals)

    # ── 阶段4: 应用变更 ──
    confirmed = [p for p in all_proposals if p.rule_id == "_FINAL_" and (p.accepted or p.user_override)]
    rejected = [p for p in all_proposals if p.rule_id == "_FINAL_" and not p.accepted and not p.user_override]

    print(f"\n  ✅ 接受 {len(confirmed)} 处, 拒绝 {len(rejected)} 处")

    if not confirmed:
        print("  没有需要应用的变更")
        return filepath

    applied = apply_confirmed_changes(wb, all_proposals)

    # 可选格式美化
    if not auto_accept:
        any_fmt = any(settings.get(f"10.0{i}_{'freeze_header' if i==1 else 'header_style' if i==2 else 'auto_column_width' if i==3 else 'auto_filter'}", {}).get("enabled")
                      for i in range(1, 5))
        if any_fmt:
            for sn in sheets:
                ws = wb[sn]
                fmt_choice = input(f"  是否对 [{sn}] 应用格式美化? (y/N): ").strip().lower()
                if fmt_choice == 'y':
                    apply_output_formatting(ws, settings)

    # ── 阶段5: 保存 ──
    dir_name = os.path.dirname(filepath) or '.'
    base = os.path.splitext(os.path.basename(filepath))[0]
    ext = os.path.splitext(filepath)[1]
    output_path = os.path.join(dir_name, f"{base}_标准化{ext}")

    # 验证保存
    wb.save(output_path)
    try:
        test_wb = load_workbook(output_path)
        test_wb.close()
    except Exception as e:
        print(f"  ⚠️  保存后验证失败: {e}")

    print(f"\n  📄 标准化文件已保存: {output_path}")

    # 导出变更日志
    # 将confirmed写入logger
    for p in all_proposals:
        if p.rule_id == "_FINAL_" and (p.accepted or p.user_override):
            logger.log(p.sheet, p.cell_ref, p.original, p.final_value, "最终变更")

    if logger.logs:
        log_path = os.path.join(dir_name, f"{base}_变更日志.xlsx")
        log_count = logger.export_to_workbook(log_path)
        print(f"  📋 变更日志已保存: {log_path} ({log_count} 条)")

    print(f"\n  ✅ 处理完成! 共应用 {applied} 处变更")
    return output_path


# ============================================================
#  终端界面
# ============================================================

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')


def print_banner():
    print("=" * 70)
    print("    Excel Standardizer v3.0 — 安全优先·两阶段处理")
    print("    核心原则: 宁可漏改,不可误改 | 公式不动 | 格式不改")
    print("=" * 70)
    print()


def print_settings_menu(settings):
    categories = OrderedDict([
        ("一、字符编码与不可见字符", [k for k in settings if k.startswith("1.")]),
        ("二、大小写(默认关闭!)", [k for k in settings if k.startswith("2.")]),
        ("三、中文相关(默认关闭)", [k for k in settings if k.startswith("3.")]),
        ("四、标点符号", [k for k in settings if k.startswith("4.")]),
        ("五、数字与数值", [k for k in settings if k.startswith("5.")]),
        ("六、日期时间(默认关闭)", [k for k in settings if k.startswith("6.")]),
        ("七、公司名称(默认关闭,需列感知)", [k for k in settings if k.startswith("7.")]),
        ("八、通用文本", [k for k in settings if k.startswith("8.")]),
        ("九、电话证件(默认关闭)", [k for k in settings if k.startswith("9.")]),
        ("十、格式美化(全部默认关闭!)", [k for k in settings if k.startswith("10.")]),
        ("十一、数据结构", [k for k in settings if k.startswith("11.")]),
        ("十二、地址(默认关闭)", [k for k in settings if k.startswith("12.")]),
    ])
    idx = 1
    key_map = {}
    for cat, keys in categories.items():
        if not keys:
            continue
        print(f"\n  ┌── {cat} ──")
        for k in keys:
            s = settings[k]
            # 安全等级标识
            level = "🟢" if k in SAFETY_LEVELS['SAFE'] else \
                    "🟡" if k in SAFETY_LEVELS['MODERATE'] else "🔴"
            status = "✅开" if s["enabled"] else "❌关"
            opt = f" [{s['option']}]" if "option" in s and s["enabled"] else ""
            # 列限制标识
            col_note = ""
            if k in RULE_COLUMN_APPLICABILITY:
                types = RULE_COLUMN_APPLICABILITY[k]
                col_note = f" ⟨仅:{','.join(types)}⟩"
            print(f"  │ {idx:3d}. {level}{status}{opt} {s['desc']}{col_note}")
            key_map[idx] = k
            idx += 1
        print(f"  └{'─' * 55}")
    print(f"\n  图例: 🟢安全 🟡中等 🔴危险(需确认列类型)")
    return key_map


def select_sheets(wb):
    print("\n  工作表列表:")
    for i, name in enumerate(wb.sheetnames, 1):
        ws = wb[name]
        print(f"    {i}. {name} ({ws.max_row}行 x {ws.max_column}列)")
    print(f"\n  输入编号(逗号分隔)或'all'处理全部:")
    choice = input("  > ").strip()
    if choice.lower() == 'all':
        return wb.sheetnames
    selected = []
    for part in choice.split(','):
        try:
            idx = int(part.strip())
            if 1 <= idx <= len(wb.sheetnames):
                selected.append(wb.sheetnames[idx - 1])
        except ValueError:
            pass
    return selected if selected else wb.sheetnames


def settings_interface(settings):
    while True:
        clear_screen()
        print_banner()
        print("  ⚙️  设置管理 (🟢安全 🟡中等 🔴危险)")
        print("-" * 70)
        key_map = print_settings_menu(settings)

        print(f"\n  操作: 编号=切换开关 | '编号 选项'=改选项 | 'safe'=仅安全规则")
        print(f"        'all_on'/'all_off' | 'save'保存 | 'back'返回")
        choice = input("\n  > ").strip()

        if choice.lower() == 'back':
            return settings
        elif choice.lower() == 'save':
            save_settings(settings)
            print("  ✅ 已保存!")
            input("  按回车继续...")
        elif choice.lower() == 'all_on':
            for k in settings:
                settings[k]["enabled"] = True
            print("  ✅ 全部开启! ⚠️  请注意危险规则可能造成误改")
            input("  按回车继续...")
        elif choice.lower() == 'all_off':
            for k in settings:
                settings[k]["enabled"] = False
            print("  ✅ 全部关闭!")
            input("  按回车继续...")
        elif choice.lower() == 'safe':
            # 只启用SAFE级别
            for k in settings:
                settings[k]["enabled"] = k in SAFETY_LEVELS['SAFE']
            # 同时启用一些MODERATE中的关键项
            for k in ["1.01_fullwidth_to_halfwidth", "1.02_fullwidth_space",
                       "1.04_tab_to_space", "1.10_unicode_normalize", "11.01_header_clean"]:
                if k in settings:
                    settings[k]["enabled"] = True
            print("  ✅ 已切换为安全模式(仅安全+部分中等规则)")
            input("  按回车继续...")
        else:
            parts = choice.split(maxsplit=1)
            try:
                num = int(parts[0])
                if num in key_map:
                    key = key_map[num]
                    if len(parts) == 1:
                        settings[key]["enabled"] = not settings[key]["enabled"]
                        st = "开启" if settings[key]["enabled"] else "关闭"
                        if settings[key]["enabled"] and key in SAFETY_LEVELS['DANGEROUS']:
                            print(f"  ⚠️  [{key}] 已{st} — 这是危险规则,请确认目标列类型!")
                        else:
                            print(f"  ✅ [{key}] 已{st}")
                    elif len(parts) == 2 and "options" in settings[key]:
                        new_opt = parts[1].strip()
                        if new_opt in settings[key]["options"]:
                            settings[key]["option"] = new_opt
                            print(f"  ✅ [{key}] 选项 → {new_opt}")
                        else:
                            print(f"  ❌ 无效选项。可选: {', '.join(settings[key]['options'])}")
                else:
                    print("  ❌ 无效编号")
            except ValueError:
                print("  ❌ 无效输入")
            input("  按回车继续...")
    return settings


def main():
    settings = load_settings()

    while True:
        clear_screen()
        print_banner()
        print("  主菜单:")
        print("-" * 70)
        print("  1. 📂 处理Excel文件(两阶段: 扫描→审核→应用)")
        print("  2. ⚙️  设置管理")
        print("  3. 💾 保存当前设置")
        print("  4. 🔄 恢复默认设置")
        print("  5. 📋 查看当前设置")
        print("  6. 🧪 生成综合测试文件")
        print("  0. 🚪 退出")
        print("-" * 70)

        choice = input("  请选择 [0-6]: ").strip()

        if choice == '0':
            print("\n  再见!")
            break
        elif choice == '1':
            print("\n  请输入Excel文件路径:")
            filepath = input("  > ").strip().strip('"').strip("'")
            if not os.path.exists(filepath):
                print(f"  ❌ 文件不存在: {filepath}")
                input("  按回车继续...")
                continue
            if not filepath.lower().endswith(('.xlsx', '.xlsm')):
                print("  ❌ 仅支持 .xlsx / .xlsm")
                input("  按回车继续...")
                continue
            try:
                process_file_v3(filepath, settings)
            except Exception as e:
                print(f"\n  ❌ 处理出错: {e}")
                import traceback
                traceback.print_exc()
            input("\n  按回车继续...")
        elif choice == '2':
            settings = settings_interface(settings)
        elif choice == '3':
            save_settings(settings)
            print("  ✅ 已保存:", SETTINGS_FILE)
            input("  按回车继续...")
        elif choice == '4':
            settings = copy.deepcopy(DEFAULT_SETTINGS)
            if os.path.exists(SETTINGS_FILE):
                os.remove(SETTINGS_FILE)
            print("  ✅ 已恢复默认!")
            input("  按回车继续...")
        elif choice == '5':
            clear_screen()
            print_banner()
            print_settings_menu(settings)
            input("\n  按回车返回...")
        elif choice == '6':
            generate_comprehensive_test_file()
            input("\n  按回车继续...")
        else:
            print("  ❌ 无效选择")
            input("  按回车继续...")


# ============================================================
#  综合测试文件生成器 (大幅增强)
# ============================================================

def generate_comprehensive_test_file():
    """生成全面的测试文件, 覆盖所有规则和边界情况"""
    wb = Workbook()

    # ═══ Sheet 1: 安全规则测试 ═══
    ws1 = wb.active
    ws1.title = "安全规则测试"
    ws1.append(["测试项", "测试输入", "预期输出", "测试的规则"])

    safe_tests = [
        ["全角ASCII", "Ｈｅｌｌｏ　Ｗｏｒｌｄ！", "Hello World!", "1.01"],
        ["全角数字", "１２３４５", "12345", "1.01"],
        ["全角空格", "你好\u3000世界", "你好 世界", "1.02"],
        ["不间断空格", "hello\xa0world\xa0test", "hello world test", "1.03"],
        ["Tab", "col1\tcol2\tcol3", "col1 col2 col3", "1.04"],
        ["零宽字符", "te\u200bst\u200cda\u200dta\ufeff", "testdata", "1.08"],
        ["控制字符", "abc\x01\x02\x03def\x7f", "abcdef", "1.09"],
        ["软连字符", "re\u00adsu\u00adme", "resume", "1.11"],
        ["双向标记", "hello\u200eworld\u200f", "helloworld", "1.12"],
        ["连续空格", "too   many     spaces", "too many spaces", "1.06"],
        ["首尾空白", "   hello world   ", "hello world", "1.07"],
        ["Unicode NFC", "e\u0301", "é", "1.10 NFC"],
        ["HTML实体&amp;", "&amp; &lt;tag&gt; &#39;", "& <tag> '", "8.07"],
        ["HTML实体&#", "&#x41;&#x42;", "AB", "8.07"],
        ["邮箱大小写", " Test@Example.COM ", "test@example.com", "8.09"],
        ["波浪号", "100～200", "100~200", "4.02"],
        ["省略号", "等等\u2026更多", "等等...更多", "4.05"],
        ["会计负数", "(1,500)", "-(1,500)", "5.06"],  # 注:千分位默认不处理
        ["表头换行", "公司\n名称", "公司 名称", "11.01"],
    ]
    for row in safe_tests:
        ws1.append(row)

    # 首行加粗
    for cell in ws1[1]:
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='C6EFCE')

    # ═══ Sheet 2: 危险规则测试(默认应被跳过) ═══
    ws2 = wb.create_sheet("危险规则测试(默认跳过)")
    ws2.append(["测试项", "测试输入", "如果规则启用会变成", "为什么默认关闭"])

    dangerous_tests = [
        ["大小写lower", "John Smith", "john smith", "人名被小写化"],
        ["大小写lower", "IBM Corporation", "ibm corporation", "缩写被小写化"],
        ["大小写lower", "README.md", "readme.md", "文件名被改变"],
        ["繁体→简体", "這個軟體很優秀", "这个软件很优秀", "港澳台用户数据被破坏"],
        ["繁体→简体", "乾隆皇帝", "干隆皇帝(歧义)", "繁简一对多歧义"],
        ["中文标点", "你好，世界！", "你好,世界!", "中文标点本身是正确的"],
        ["中文顿号", "甲、乙、丙", "甲,乙,丙", "法律文书顿号有语义"],
        ["括号统一旧", "data[0] = {key: val}", "data(0) = (key: val)", "代码被破坏"],
        ["括号统一旧", "《红楼梦》【精装】", "(红楼梦)(精装)", "书名号被破坏"],
        ["千分位旧", "件号: A1,234B", "件号: A1234B", "非数字逗号被删"],
        ["货币符号", "$100 vs ¥200", "100 vs 200", "币种信息丢失"],
        ["公司后缀旧", "I have limited time", "I have ltd time", "普通文本被误改"],
        ["公司停用词旧", "The University of Oxford", "University Oxford", "机构名被破坏"],
        ["空值旧", "-", "(空)", "减号/分隔符被清空"],
        ["空值旧", "—", "(空)", "破折号被当空值"],
        ["布尔值", "1", "Yes", "编号1被当布尔值"],
        ["省市区后缀", "海淀区", "海淀", "可能造成歧义"],
        ["中文数字", "三十里铺", "30里铺", "地名被破坏"],
    ]
    for row in dangerous_tests:
        ws2.append(row)

    for cell in ws2[1]:
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='FFC7CE')

    # ═══ Sheet 3: 公式保护测试 ═══
    ws3 = wb.create_sheet("公式保护测试")
    ws3.append(["描述", "数值A", "数值B", "公式(不应被改)"])
    ws3.append(["加法", 100, 200, "=B2+C2"])
    ws3.append(["求和", 300, 400, "=SUM(B2:C3)"])
    ws3.append(["条件", "Yes", "No", '=IF(B4="Yes","通过","未通过")'])
    ws3.append(["文本连接", "Hello", "World", "=B5&C5"])
    ws3.append(["VLOOKUP", 1, 2, "=VLOOKUP(B6,B2:C5,2,FALSE)"])

    for cell in ws3[1]:
        cell.font = openpyxl.styles.Font(bold=True, size=11)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='BDD7EE')

    # ═══ Sheet 4: 数字精度测试 ═══
    ws4 = wb.create_sheet("数字精度测试")
    ws4.append(["描述", "值(勿改)", "备注"])
    ws4.append(["整数", 12345, "应保持不变"])
    ws4.append(["浮点数", 3.14159265358979, "应保持不变"])
    ws4.append(["大数(float)", 1234567890123456.0, "可能被Excel截断,应警告"])
    ws4.append(["文本数字(身份证)", "310101200001011234", "文本格式,18位不应被截断"])
    ws4.append(["科学计数法文本", "1.23E+06", "默认不转换(关闭状态)"])
    ws4.append(["千分位文本", "1,234,567.89", "默认不处理千分位"])
    ws4.append(["负会计格式", "(500)", "-500 (默认处理)"])
    ws4.append(["百分比", "85.5%", "默认不处理"])
    ws4.append(["货币文本", "$1,234.56", "默认不处理"])
    ws4.append(["日期对象", datetime(2024, 3, 15), "日期对象应保持不变"])
    ws4.append(["布尔True", True, "布尔应保持不变"])
    ws4.append(["布尔False", False, "布尔应保持不变"])

    for cell in ws4[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='FCE4D6')

    # ═══ Sheet 5: 综合混合数据(列类型感知测试) ═══
    ws5 = wb.create_sheet("综合混合数据")
    ws5.append(["姓名", "公司名称", "联系电话", "邮箱", "金额", "地址", "备注", "身份证号"])

    mixed = [
        ["　張三　", "華為技術（深圳）有限責任公司", "+86-138-0000-1234",
         " Zhangsan@GMAIL.COM ", "￥1,234,567.89", "北京市朝陽區xx路", "N/A",
         "110101199001011234"],
        [" Ｊｏｈｎ　Ｄｏｅ ", "Apple\xa0Inc.\u200b", "(1) 555-123-4567",
         "john.DOE@apple.com", "$2,500,000", "123 Main St., New York", "none",
         "310101200001011234"],
        ["田中太郎", "株式会社ソフトバンク", "03-1234-5678",
         "tanaka@softbank.co.jp", "¥1,000,000", "東京都渋谷区", "　暂无　", ""],
        ["이민호", "Samsung\tElectronics Co.,\nLtd.", "010-1234-5678",
         "LEE@Samsung.COM", "₩5,000,000", "Seoul, Korea", "NULL", ""],
        ["Maria García", "Banco Santander，S.A.", "+34-91-123-4567",
         "maria@santander.es", "€1.500", "Madrid，Spain", "—", ""],
        ["Müller", "Volkswagen\u200bAktiengesellschaft", "+49 30 1234 5678",
         "Mueller@VW.DE", "€500,000", "Berlin, Germany", "-", ""],
        ["　李四　", "腾讯科技（深圳）有限责任公司", "0755-86013388",
         "lisi@tencent.com", "RMB 999,999.99", "深圳市南山区科技园", "无", "44030119850101123x"],
        ["O'Brien", "The Goldman Sachs Group, Inc.", "+1-212-555-0199",
         "obrien@gs.COM", "US$10,000,000", "200 West St, New York", "---", ""],
    ]
    for row in mixed:
        ws5.append(row)

    for cell in ws5[1]:
        cell.font = openpyxl.styles.Font(bold=True, size=12)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='D9E1F2')

    # 合并单元格测试
    ws5.merge_cells('A11:H11')
    ws5['A11'] = '以上为测试数据 — 此行为合并单元格(不应被破坏)'
    ws5['A11'].font = openpyxl.styles.Font(italic=True, color='808080')
    ws5['A11'].alignment = openpyxl.styles.Alignment(horizontal='center')

    # ═══ Sheet 6: 特殊字符边界情况 ═══
    ws6 = wb.create_sheet("特殊字符边界")
    ws6.append(["类型", "测试值", "说明"])

    edge_cases = [
        ["日文片假名", "アイウエオ", "全角片假名不应被转换为半角"],
        ["日文混合", "東京タワー（Tokyo Tower）", "日文+英文混合"],
        ["韩文", "삼성전자 주식회사", "韩文不应被破坏"],
        ["Box Drawing", "┌─────┐\n│  OK  │\n└─────┘", "表格线字符不应被Emoji清除"],
        ["数学符号", "∑ ∫ ± × ÷ √ ∞ ≠ ≈ π", "数学符号不应被清除"],
        ["带括号代码", "array[0] = {key: 'val'}", "[]和{}不应被转为()"],
        ["JSON", '{"name": "test", "arr": [1,2,3]}', "JSON不应被改"],
        ["邮箱vs@符号", "@username (not email)", "不是邮箱不应被处理"],
        ["路径", "C:\\Users\\admin\\file.txt", "文件路径"],
        ["正则表达式", "^\\d{3}-\\d{4}$", "正则不应被改"],
        ["Unicode组合", "café naïve résumé", "NFC只合并,不丢变音"],
        ["上标下标", "H₂O CO² 10³", "NFKC会破坏,NFC不会"],
        ["连字", "ﬁﬂﬀﬃ", "NFC保留,NFKC会拆开"],
        ["零宽空格混合", "hel\u200blo\u200c wo\u200drld", "零宽清除但空格保留"],
        ["多种空白混合", "\t hello\xa0\u3000world\r\n  test  ", "各类空白统一处理"],
        ["纯数字文本", "0001234567890", "文本格式数字不应丢前导零"],
        ["空字符串", "", "空值不应报错"],
        ["超长文本", "A" * 500, "超长文本不应截断"],
        ["特殊HTML", "&lt;script&gt;alert('xss')&lt;/script&gt;",
         "<script>alert('xss')</script>"],
        ["中英混排", "张三(Zhang San)于2024年1月1日签约",
         "中英混排的标准化"],
    ]
    for row in edge_cases:
        ws6.append(row)

    for cell in ws6[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='E2EFDA')

    # ═══ Sheet 7: 格式保留测试 ═══
    ws7 = wb.create_sheet("格式保留测试")
    ws7.append(["内容", "应保留的格式"])

    test_formats = [
        (" hello world ", "红色粗体"),
        ("　全角空格　", "蓝色斜体"),
        ("test\xa0nbsp\xa0data", "黄色背景"),
        ("zero\u200bwidth", "下划线"),
        ("ctrl\x01char\x02test", "大字号20pt"),
        ("  spaces  galore  ", "居中对齐"),
    ]

    colors = ['FF0000', '0000FF', '000000', '000000', '000000', '000000']
    bolds = [True, False, False, False, False, False]
    italics = [False, True, False, False, False, False]
    fills_hex = [None, None, 'FFFF00', None, None, None]
    underlines = [None, None, None, 'single', None, None]
    sizes = [11, 11, 11, 11, 20, 11]
    aligns = [None, None, None, None, None, 'center']

    for i, (content, desc) in enumerate(test_formats):
        r = i + 2
        c1 = ws7.cell(row=r, column=1, value=content)
        c2 = ws7.cell(row=r, column=2, value=desc)

        c1.font = openpyxl.styles.Font(
            bold=bolds[i], italic=italics[i], color=colors[i],
            size=sizes[i], underline=underlines[i]
        )
        if fills_hex[i]:
            c1.fill = openpyxl.styles.PatternFill('solid', fgColor=fills_hex[i])
        if aligns[i]:
            c1.alignment = openpyxl.styles.Alignment(horizontal=aligns[i])

    for cell in ws7[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='D9D9D9')

    # ═══ Sheet 8: 日期时间测试 ═══
    ws8 = wb.create_sheet("日期时间测试")
    ws8.append(["测试项", "输入", "预期(YYYY-MM-DD)", "说明"])

    date_tests = [
        ["标准斜线", "2024/01/15", "2024-01-15", ""],
        ["标准点号", "2024.3.5", "2024-03-05", ""],
        ["中文日期", "2024年1月15日", "2024-01-15", ""],
        ["英文月份", "Jan 15, 2024", "2024-01-15", ""],
        ["英文月份2", "15 March 2024", "2024-03-15", ""],
        ["欧式DMY", "25/12/2024", "2024-12-25", "日>12推断为DMY"],
        ["歧义MDY", "01/02/2024", "2024-01-02", "默认美式MDY"],
        ["12小时制", "3:05 PM", "15:05:00", "时间"],
        ["24小时制", "15:05:30", "15:05:30", "时间"],
        ["非日期", "hello world", "hello world", "不应被改"],
        ["日期对象", None, None, "openpyxl日期对象"],
    ]
    for row in date_tests:
        ws8.append(row)
    # 写入一个真正的日期对象
    ws8.cell(row=12, column=2, value=datetime(2024, 6, 15))
    ws8.cell(row=12, column=3, value="保持datetime对象")

    for cell in ws8[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # ═══ Sheet 9: 公司名称测试 ═══
    ws9 = wb.create_sheet("公司名称测试")
    ws9.append(["公司名称", "说明"])

    company_tests = [
        ["Apple Inc.", "英文公司"],
        ["Google LLC", "LLC后缀"],
        ["Samsung Electronics Co., Ltd.", "多后缀"],
        ["Huawei Technologies Co., Limited", "含tech"],
        ["中国平安保险（集团）股份有限公司", "中文公司"],
        ["腾讯科技（深圳）有限责任公司", "有限责任→有限"],
        ["The Goldman Sachs Group, Inc.", "含停用词the"],
        ["Deutsche Bank AG", "国际后缀AG"],
        ["Toyota Motor Corporation", "Corporation"],
        ["PT Telkom Indonesia Tbk", "印尼后缀"],
        ["I have limited time to finish this", "不是公司名!不应被改"],
        ["The service is incorporated into the system", "不是公司名!"],
        ["Technology is amazing", "不是公司名!"],
    ]
    for row in company_tests:
        ws9.append(row)

    for cell in ws9[1]:
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill('solid', fgColor='E2EFDA')

    # 保存
    output_path = os.path.join(os.getcwd(), "test_standardization_v3.xlsx")
    wb.save(output_path)
    print(f"\n  ✅ 综合测试文件已生成: {output_path}")
    print("  包含9个Sheet:")
    print("    1. 安全规则测试 - 测试默认启用的安全规则")
    print("    2. 危险规则测试 - 验证默认关闭的危险规则不会触发")
    print("    3. 公式保护测试 - 验证公式不被修改")
    print("    4. 数字精度测试 - 验证数字/日期/布尔不被改")
    print("    5. 综合混合数据 - 模拟真实业务数据(列类型感知)")
    print("    6. 特殊字符边界 - 各种边界情况和特殊字符")
    print("    7. 格式保留测试 - 验证字体/颜色/背景不被覆盖")
    print("    8. 日期时间测试 - 各种日期格式")
    print("    9. 公司名称测试 - 公司名vs普通文本区分")
    return output_path


# ============================================================
#  CLI入口
# ============================================================

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == '--test':
        generate_comprehensive_test_file()
    elif len(sys.argv) > 1 and sys.argv[1] == '--auto':
        if len(sys.argv) < 3:
            print("用法: python excel_standardizer.py --auto <文件路径>")
            sys.exit(1)
        filepath = sys.argv[2]
        settings = load_settings()
        process_file_v3(filepath, settings, auto_accept=True)
    else:
        main()